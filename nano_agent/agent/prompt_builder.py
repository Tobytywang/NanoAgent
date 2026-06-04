"""
模块化 Prompt 构建器

支持:
1. 从 Excel 配置文件加载模块配置
2. 按需组装 prompt
3. Token 预算控制
4. 动态模块激活（基于意图检测）
"""

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .prompt_modules import MODULES, STYLE_PRESETS, PromptModule
from .intent_detector import IntentDetector


@dataclass
class PromptBuilderConfig:
    """Prompt 构建器配置"""

    style: str = "standard"
    modules: list[str] = field(default_factory=list)
    token_budget: int = 2000
    include_environment: bool = False
    include_git: bool = False
    custom_modules: dict[str, PromptModule] = field(default_factory=dict)
    # Dynamic module activation
    dynamic_module_activation: bool = True  # Enable intent-based module activation


class ExcelConfigManager:
    """
    Excel 配置管理器

    用于读取和写入 Prompt 模块配置到 Excel 文件。
    """

    # Excel 列定义
    COLUMNS = [
        ("name", "模块名称"),
        ("description", "描述"),
        ("priority", "优先级"),
        ("always_on", "始终启用"),
        ("enabled", "当前启用"),
        ("token_estimate", "预估Tokens"),
        ("is_stable", "是否稳定"),
        ("category", "模块分类"),
        ("content", "内容模板"),
    ]

    # Style 配置列
    STYLE_COLUMNS = [
        ("style_name", "风格名称"),
        ("description", "描述"),
        ("modules", "包含模块(逗号分隔)"),
        ("token_budget", "Token预算"),
    ]

    def __init__(self, config_path: str | Path | None = None):
        """
        初始化 Excel 配置管理器

        Args:
            config_path: Excel 配置文件路径，默认为 nano_agent/agent/prompts.xlsx
        """
        if config_path is None:
            config_path = Path(__file__).parent / "prompts.xlsx"
        self.config_path = Path(config_path)

    def load(self) -> dict[str, PromptModule]:
        """
        从 Excel 加载模块配置

        Returns:
            模块名称到 PromptModule 的映射
        """
        if not self.config_path.exists():
            return {}

        wb = load_workbook(self.config_path)
        modules = {}

        # 读取模块配置表
        if "Modules" in wb.sheetnames:
            ws = wb["Modules"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                name = row[0]
                modules[name] = PromptModule(
                    name=name,
                    description=row[1] or "",
                    priority=int(row[2]) if row[2] else 50,
                    always_on=bool(row[3]) if row[3] else False,
                    enabled=bool(row[4]) if row[4] else True,
                    token_estimate=int(row[5]) if row[5] else 0,
                    is_stable=bool(row[6]) if row[6] else True,
                    category=row[7] or "core" if len(row) > 7 else "core",
                    content=row[8] or "" if len(row) > 8 else (row[6] or ""),
                )

        return modules

    def load_styles(self) -> dict[str, dict]:
        """
        从 Excel 加载 Style 配置

        Returns:
            Style 名称到配置的映射
        """
        if not self.config_path.exists():
            return {}

        wb = load_workbook(self.config_path)
        styles = {}

        if "Styles" in wb.sheetnames:
            ws = wb["Styles"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                style_name = row[0]
                modules_str = row[2] or ""
                styles[style_name] = {
                    "description": row[1] or "",
                    "modules": [m.strip() for m in modules_str.split(",") if m.strip()],
                    "token_budget": int(row[3]) if row[3] else 2000,
                }

        return styles

    def save(
        self, modules: dict[str, PromptModule] | None = None, styles: dict | None = None
    ):
        """
        保存配置到 Excel

        Args:
            modules: 模块配置，默认使用预定义模块
            styles: Style 配置，默认使用预定义配置
        """
        if modules is None:
            modules = MODULES
        if styles is None:
            styles = STYLE_PRESETS

        wb = Workbook()

        # 创建模块配置表
        ws_modules = wb.active
        ws_modules.title = "Modules"

        # 写入表头
        for col_idx, (key, header) in enumerate(self.COLUMNS, 1):
            ws_modules.cell(row=1, column=col_idx, value=header)

        # 写入模块数据
        for row_idx, (name, module) in enumerate(modules.items(), 2):
            ws_modules.cell(row=row_idx, column=1, value=name)
            ws_modules.cell(row=row_idx, column=2, value=module.description)
            ws_modules.cell(row=row_idx, column=3, value=module.priority)
            ws_modules.cell(row=row_idx, column=4, value=module.always_on)
            ws_modules.cell(row=row_idx, column=5, value=module.enabled)
            ws_modules.cell(row=row_idx, column=6, value=module.token_estimate)
            ws_modules.cell(row=row_idx, column=7, value=module.is_stable)
            ws_modules.cell(row=row_idx, column=8, value=module.category)
            ws_modules.cell(row=row_idx, column=9, value=module.content)

        # 创建 Style 配置表
        ws_styles = wb.create_sheet("Styles")

        for col_idx, (key, header) in enumerate(self.STYLE_COLUMNS, 1):
            ws_styles.cell(row=1, column=col_idx, value=header)

        for row_idx, (style_name, config) in enumerate(styles.items(), 2):
            ws_styles.cell(row=row_idx, column=1, value=style_name)
            ws_styles.cell(row=row_idx, column=2, value=config["description"])
            ws_styles.cell(row=row_idx, column=3, value=",".join(config["modules"]))
            ws_styles.cell(row=row_idx, column=4, value=config["token_budget"])

        # 创建说明表
        ws_readme = wb.create_sheet("README")
        readme_content = [
            ["Prompt 模块配置说明"],
            [""],
            ["Modules 表:"],
            ["- name: 模块唯一标识"],
            ["- description: 模块功能描述"],
            ["- priority: 优先级(数值越小越靠前)"],
            ["- always_on: 是否始终启用(不可关闭)"],
            ["- enabled: 当前是否启用"],
            ["- token_estimate: 预估 token 数量"],
            ["- is_stable: 是否属于稳定部分(适合缓存)"],
            ["- category: 模块分类(core/efficiency/security/output/context/memory)"],
            ["- content: 模块内容模板"],
            [""],
            ["Styles 表:"],
            ["- style_name: 风格名称(concise/standard/detailed)"],
            ["- modules: 包含的模块列表(逗号分隔)"],
            ["- token_budget: Token 预算上限"],
            [""],
            ["使用方法:"],
            ["1. 修改 enabled 列来启用/禁用模块"],
            ["2. 修改 Styles 表中的 modules 列来调整各风格包含的模块"],
            ["3. 修改 content 列来调整模块内容"],
            ["4. 修改 priority 列来调整模块顺序"],
        ]
        for row_idx, row_data in enumerate(readme_content, 1):
            ws_readme.cell(row=row_idx, column=1, value=row_data[0] if row_data else "")

        # 确保目录存在
        self.config_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(self.config_path)
        print(f"[PromptConfig] 配置已保存到 {self.config_path}")

    def create_default(self):
        """创建默认配置文件"""
        self.save()


class PromptBuilder:
    """
    模块化 Prompt 构建器

    支持从 Excel 配置或代码配置组装 prompt。

    Usage:
        # 使用默认配置
        builder = PromptBuilder()
        prompt = builder.build_for_style("standard", tools_description)

        # 使用 Excel 配置
        builder = PromptBuilder.from_excel("config/prompts.xlsx")
        prompt = builder.build_for_style("standard", tools_description)

        # 自定义组装
        builder = PromptBuilder()
        prompt = builder.add_module("core").add_module("security").build(tools_description)
    """

    def __init__(self, config: PromptBuilderConfig | None = None):
        self.config = config or PromptBuilderConfig()
        self._modules: dict[str, PromptModule] = {}
        self._dynamic_content: dict[str, str] = {}
        self._intent_detector = (
            IntentDetector()
        )  # Intent detector for dynamic activation
        self._load_modules()

    def _load_modules(self):
        """加载模块配置"""
        # 加载预定义模块
        self._modules = dict(MODULES)

        # 加载自定义模块
        for name, module in self.config.custom_modules.items():
            self._modules[name] = module

    @classmethod
    def from_excel(cls, config_path: str | Path) -> "PromptBuilder":
        """
        从 Excel 配置创建 Builder

        Args:
            config_path: Excel 配置文件路径

        Returns:
            PromptBuilder 实例
        """
        manager = ExcelConfigManager(config_path)
        modules = manager.load()
        styles = manager.load_styles()

        # 合并预定义模块和 Excel 模块
        all_modules = dict(MODULES)
        all_modules.update(modules)

        config = PromptBuilderConfig(
            custom_modules=all_modules,
        )

        # 如果有 styles 配置，使用第一个作为默认
        if styles:
            default_style = list(styles.keys())[0]
            config.style = default_style
            config.modules = styles[default_style].get("modules", [])
            config.token_budget = styles[default_style].get("token_budget", 2000)

        builder = cls(config)
        builder._modules = all_modules
        builder._style_presets = styles

        return builder

    def add_module(self, name: str) -> "PromptBuilder":
        """
        添加模块

        Args:
            name: 模块名称

        Returns:
            self (支持链式调用)
        """
        if name in self._modules:
            self.config.modules.append(name)
        return self

    def remove_module(self, name: str) -> "PromptBuilder":
        """
        移除模块

        Args:
            name: 模块名称

        Returns:
            self
        """
        if name in self.config.modules:
            self.config.modules.remove(name)
        return self

    def set_style(self, style: str) -> "PromptBuilder":
        """
        设置风格

        Args:
            style: concise / standard / detailed

        Returns:
            self
        """
        self.config.style = style
        if style in STYLE_PRESETS:
            self.config.modules = STYLE_PRESETS[style]["modules"].copy()
            self.config.token_budget = STYLE_PRESETS[style]["token_budget"]
        return self

    def enable_environment(self) -> "PromptBuilder":
        """启用环境信息"""
        self.config.include_environment = True
        return self

    def enable_git(self) -> "PromptBuilder":
        """启用 Git 状态"""
        self.config.include_git = True
        return self

    def _get_environment_content(self) -> str:
        """获取环境信息"""
        import platform

        return f"""## Environment
- Working directory: {os.getcwd()}
- Platform: {platform.system()}
- Python: {platform.python_version()}
"""

    def _get_git_content(self) -> str:
        """获取 Git 状态"""
        try:
            import subprocess

            branch = (
                subprocess.check_output(
                    ["git", "branch", "--show-current"], stderr=subprocess.DEVNULL
                )
                .decode()
                .strip()
            )

            status = (
                subprocess.check_output(
                    ["git", "status", "--short"], stderr=subprocess.DEVNULL
                )
                .decode()
                .strip()[:100]
            )

            return f"""## Git Status
- Branch: {branch}
- Changes: {status or 'clean'}
"""
        except Exception:
            return ""

    def build(self, tools_description: str = "") -> str:
        """
        构建 prompt

        Args:
            tools_description: 工具描述

        Returns:
            完整的 system prompt
        """
        parts = []

        # 按优先级排序模块
        enabled_modules = [
            self._modules[name]
            for name in self.config.modules
            if name in self._modules and self._modules[name].enabled
        ]
        enabled_modules.sort(key=lambda m: m.priority)

        # 组装各模块
        for module in enabled_modules:
            content = module.content

            # 替换占位符
            if "{tools_description}" in content:
                content = content.replace("{tools_description}", tools_description)

            parts.append(content)

        # 添加动态内容
        if self.config.include_environment:
            parts.append(self._get_environment_content())

        if self.config.include_git:
            parts.append(self._get_git_content())

        return "\n\n".join(parts)

    def build_for_style(self, style: str, tools_description: str = "") -> str:
        """
        按风格构建 prompt

        Args:
            style: concise / standard / detailed
            tools_description: 工具描述

        Returns:
            完整的 system prompt
        """
        self.set_style(style)
        return self.build(tools_description)

    def estimate_tokens(self) -> int:
        """估算 token 数量"""
        total = 0
        for name in self.config.modules:
            if name in self._modules:
                total += self._modules[name].effective_token_estimate

        if self.config.include_environment:
            total += 50
        if self.config.include_git:
            total += 80

        return total

    def get_module_info(self) -> list[dict]:
        """获取模块信息列表"""
        return [
            {
                "name": m.name,
                "description": m.description,
                "priority": m.priority,
                "enabled": m.enabled,
                "tokens": m.effective_token_estimate,
                "is_stable": m.is_stable,
                "category": m.category,
            }
            for m in sorted(self._modules.values(), key=lambda x: x.priority)
        ]

    def build_stable(
        self, tools_description: str = "", stable_modules: list[str] | None = None
    ) -> str:
        """
        构建稳定部分（不频繁变化，适合 LLM API 缓存）

        Args:
            tools_description: 工具描述
            stable_modules: 稳定模块列表，默认使用配置中的 stable_modules

        Returns:
            稳定部分的 system prompt
        """
        if stable_modules is None:
            stable_modules = self.config.modules

        # 筛选稳定模块
        stable_module_names = [
            name
            for name in stable_modules
            if name in self._modules
            and self._modules[name].enabled
            and self._modules[name].is_stable
        ]

        # 按优先级排序
        stable_modules_list = [self._modules[name] for name in stable_module_names]
        stable_modules_list.sort(key=lambda m: m.priority)

        parts = []
        for module in stable_modules_list:
            content = module.content
            if "{tools_description}" in content:
                content = content.replace("{tools_description}", tools_description)
            parts.append(content)

        return "\n\n".join(parts)

    def build_dynamic(
        self,
        skill_prompt: str = "",
        confidence_enabled: bool = False,
        confidence_suffix: str = "",
        user_input: str | None = None,
    ) -> str:
        """
        构建动态部分（每次可能变化）

        Args:
            skill_prompt: 技能 prompt
            confidence_enabled: 是否启用置信度
            confidence_suffix: 置信度后缀内容
            user_input: 用户输入（用于意图检测，决定是否激活 git/environment 模块）

        Returns:
            动态部分的 system prompt
        """
        parts = []

        # 添加动态模块（is_stable=False）
        dynamic_modules = [
            self._modules[name]
            for name in self.config.modules
            if name in self._modules
            and self._modules[name].enabled
            and not self._modules[name].is_stable
        ]
        dynamic_modules.sort(key=lambda m: m.priority)

        for module in dynamic_modules:
            if module.content:
                parts.append(module.content)

        # 添加技能 prompt
        if skill_prompt:
            parts.append(f"## Skills\n\n{skill_prompt}")

        # 添加置信度后缀
        if confidence_enabled and confidence_suffix:
            parts.append(confidence_suffix)

        # 动态模块激活：基于意图检测决定是否添加 environment/git
        should_include_env = self.config.include_environment
        should_include_git = self.config.include_git

        # 如果启用了动态激活且有用户输入，则进行意图检测
        if self.config.dynamic_module_activation and user_input:
            detected_intents = self._intent_detector.detect(user_input)
            # 只有当意图检测到且配置未显式禁用时才激活
            if "environment" in detected_intents:
                should_include_env = True
            if "git_status" in detected_intents:
                should_include_git = True

        # 添加环境信息
        if should_include_env:
            parts.append(self._get_environment_content())

        # 添加 Git 状态
        if should_include_git:
            parts.append(self._get_git_content())

        return "\n\n".join(parts)

    def get_cache_key(self, tools_description: str = "") -> str:
        """
        获取稳定部分的缓存键（用于 LLM API 缓存优化）

        Args:
            tools_description: 工具描述

        Returns:
            缓存键字符串
        """
        import hashlib

        stable_prompt = self.build_stable(tools_description)
        # Use SHA256 for stable cache keys across sessions
        hash_bytes = hashlib.sha256(stable_prompt.encode()).digest()
        return f"prompt_{hash_bytes[:4].hex()}"

    def get_stable_module_names(self) -> list[str]:
        """获取稳定模块名称列表"""
        return [
            name
            for name in self.config.modules
            if name in self._modules
            and self._modules[name].enabled
            and self._modules[name].is_stable
        ]

    def get_dynamic_module_names(self) -> list[str]:
        """获取动态模块名称列表"""
        return [
            name
            for name in self.config.modules
            if name in self._modules
            and self._modules[name].enabled
            and not self._modules[name].is_stable
        ]

    def get_intent_detector(self) -> IntentDetector:
        """获取意图检测器"""
        return self._intent_detector

    def detect_intents(self, user_input: str) -> set[str]:
        """检测用户意图

        Args:
            user_input: 用户输入

        Returns:
            检测到的意图集合
        """
        return self._intent_detector.detect(user_input)

    # ── Tool description formatting ────────────────────────────────────

    @staticmethod
    def format_tools_description(tool_registry, style: str = "standard") -> str:
        """Format tool descriptions from the registry.

        Args:
            tool_registry: ToolRegistry instance
            style: Output style - "concise", "standard", or "detailed"

        Returns:
            Formatted tool description string
        """
        from .prompts import TOOL_DESCRIPTION_TEMPLATE

        tools = tool_registry.list_tools()

        if style == "concise":
            # Only tool names, comma separated (minimal tokens)
            return ", ".join(tools)

        descriptions = []
        for name in tools:
            tool = tool_registry.get(name)

            if style == "standard":
                first_sentence = tool.description.split(".")[0]
                required = tool.parameters_schema.get("required", [])
                params_str = ", ".join(required) if required else "none"
                descriptions.append(
                    f"- {name}: {first_sentence}\n  params: {params_str}"
                )
            else:
                desc = TOOL_DESCRIPTION_TEMPLATE.format(
                    name=name,
                    description=tool.description,
                    parameters=tool.parameters_schema,
                )
                descriptions.append(desc)

        return "\n".join(descriptions)


# 便捷函数
def build_prompt(
    tools_description: str,
    style: str = "standard",
    include_env: bool = False,
    include_git: bool = False,
    config_path: str | Path | None = None,
) -> str:
    """
    构建 system prompt

    Args:
        tools_description: 工具描述
        style: concise / standard / detailed
        include_env: 是否包含环境信息
        include_git: 是否包含 Git 状态
        config_path: Excel 配置路径（可选）

    Returns:
        完整的 system prompt
    """
    if config_path:
        builder = PromptBuilder.from_excel(config_path)
    else:
        builder = PromptBuilder()

    builder.set_style(style)

    if include_env:
        builder.enable_environment()
    if include_git:
        builder.enable_git()

    return builder.build(tools_description)


def create_default_config(config_path: str | Path | None = None):
    """
    创建默认 Excel 配置文件

    Args:
        config_path: 配置文件路径，默认为 config/prompts.xlsx
    """
    manager = ExcelConfigManager(config_path)
    manager.create_default()
